import openpyxl
import re
import sys

def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False

def process_string(value):
    '''We process the string to remove ` symbol because cmd treats ` as command and not string'''
    value = value.replace("`", "'")
    return value

args_len = len(sys.argv)
if args_len == 2:
    wb = openpyxl.load_workbook(sys.argv[1])
else:
    wb = openpyxl.load_workbook('NoPhp.xlsx')

ws = wb[wb.active.title]
row_len = ws.max_row
col_len = ws.max_column


fileds = [ 'Summary', 'Package', 'Estimate', 'Blocker', 'User', 'Blocking', 'bugid', 'Description' ]
bug_args = [ '--summary', '--package', '--time', '--user', '--blocks', '--description' ]

for field, col in zip( fileds, ws.iter_cols( min_row=1, max_col=col_len, max_row=1 ) ):
    for cell in col:
        match = re.search( cell.value, field, re.IGNORECASE)
        assert match, ('Filed `' + field + '` not matching')

#Delete unwanted rows `Blocker` and `bugid` as we don't need them while creating a bug
ws.delete_cols(4)
ws.delete_cols(6)

ws.cell( row=1, column=col_len+1, value='Create Command' )
ws.column_dimensions['I'].width = 50

for idx, row in enumerate( ws.iter_rows(min_row=2, max_col=col_len, max_row=row_len ), start = 2 ):
    if row[0].value is None:
        break
    create_bug_cmd = 'a bug create '
    for cell, arg_symbol in zip( row, bug_args ):
        value = cell.value
        if value is not None:
            if isfloat(value):
                value = int(value)
            value = str(value)
            value = process_string(value)
            create_bug_cmd += ( arg_symbol + ' "' + value + '" ' )
    
    print(idx-1, create_bug_cmd)
    print()
    ws.cell( row=idx, column=col_len+1, value=create_bug_cmd )

wb.save('bugs.xlsx')

print('Please check "bugs.xlsx to find a bug create commands and review before executing!')

# Sample
# a bug create --summary "IS-IS SR Node Segment no-php flag support: Tac Model changes and Cli Plugin changes" --package "Isis-task" --time "12" --user "ashishb" --blocks "671433" --description "Add new Tac attribute noPhp and modify the existing CLI command 'node-segment ipv4|ipv6 index <val> [ explicit-null ]' with new flag no-php option. This bug also covers the required CLI test"

# Output
# opened bug 694150, assigned to ashishb@arista.com